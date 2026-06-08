import json, sys, os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

input_path = r"C:\Сделать\Чейчер SCRUM\openclaw\matrix-council-frameworks.json"
output_path = r"C:\Сделать\Чейчер SCRUM\openclaw\Таблица выбора агента для принятия решений.xlsx"

with open(input_path, encoding="utf-8") as f:
    data = json.load(f)

params = data["params"]
services = data["services"]

# Calculate usefulness for sorting
for svc in services:
    svc["_usefulness"] = sum(s * p["weight"] / 100 for s, p in zip(svc["scores"], params))

# Sort by usefulness DESC
services.sort(key=lambda x: -x["_usefulness"])

wb = Workbook()

# ============ SHEET 1: Матрица ============
ws = wb.active
ws.title = "Матрица"

hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="2F5496")
sub_fill = PatternFill("solid", fgColor="D6E4F0")
gold_fill = PatternFill("solid", fgColor="FFF2CC")
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
wrap = Alignment(wrap_text=True, vertical='center', horizontal='center')
wrap_left = Alignment(wrap_text=True, vertical='center')

# Row 1: Title
ncols = 4 + len(params)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
ws.cell(1, 1, data["title"]).font = Font(bold=True, size=14)
ws.row_dimensions[1].height = 30

# Row 2: Subtitle
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
ws.cell(2, 1, data["subtitle"]).font = Font(italic=True, size=9, color="666666")

# Row 4: Headers
headers = ["Фреймворк", "Полезность", "Цена", "Ссылка"]
for p in params:
    arrow = "↓" if p["reverse"] else "↑"
    headers.append(f"{p['name']} {arrow} ({p['weight']}%)")

for c, h in enumerate(headers, 1):
    cell = ws.cell(4, c, h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = wrap
    cell.border = thin

# Row 5: Weights
weights_row = ["ВЕС →", "", "", ""]
for p in params:
    weights_row.append(f"{p['weight']}%")
for c, w in enumerate(weights_row, 1):
    cell = ws.cell(5, c, w)
    cell.fill = sub_fill
    cell.border = thin
    cell.font = Font(bold=True, size=9)
    cell.alignment = wrap

# Data rows (sorted by usefulness)
data_start = 6
for si, svc in enumerate(services):
    row = data_start + si
    u = svc["_usefulness"]
    price = svc["price"]
    
    values = [svc["name"], round(u, 2), price if price else "Free", svc["link"]]
    for s in svc["scores"]:
        values.append(s)
    
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.border = thin
        if c == 1:
            cell.font = Font(bold=True, size=11)
            cell.alignment = wrap_left
        elif c == 4:
            cell.font = Font(color="0563C1", underline="single", size=9)
            cell.alignment = wrap_left
        else:
            cell.alignment = wrap
    
    # Highlight top 3
    if si < 3:
        medals = ["🥇", "🥈", "🥉"]
        ws.cell(row, 1).value = medals[si] + " " + svc["name"]
        ws.cell(row, 1).fill = gold_fill

data_end = data_start + len(services) - 1

# Auto-filter on header row
ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{data_end}"

# Color scales for numeric columns (2=usefulness, 5..=scores)
for col_idx in [2] + list(range(5, 5 + len(params))):
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{data_start}:{col_letter}{data_end}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',
        mid_type='num', mid_value=5, mid_color='FFEB84',
        end_type='num', end_value=10, end_color='63BE7B'
    ))

# Column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 13
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 35
for i in range(len(params)):
    ws.column_dimensions[get_column_letter(5 + i)].width = 16

# Freeze panes
ws.freeze_panes = 'A6'

# ============ SHEET 2: Обоснования ============
ws2 = wb.create_sheet("Обоснования")

ws2.cell(1, 1, "Обоснования баллов").font = Font(bold=True, size=14)
ws2.merge_cells("A1:E1")

# Headers
ws2.cell(3, 1, "Фреймворк").font = Font(bold=True)
for pi, p in enumerate(params):
    ws2.cell(3, 2 + pi, f"{p['name']} ({p['weight']}%)").font = Font(bold=True)
    ws2.cell(3, 2 + pi).alignment = wrap

for si, svc in enumerate(services):
    row = 4 + si
    ws2.cell(row, 1, svc["name"]).font = Font(bold=True)
    ws2.cell(row, 1).border = thin
    for pi, desc in enumerate(svc["descs"]):
        cell = ws2.cell(row, 2 + pi, desc)
        cell.alignment = wrap
        cell.border = thin

ws2.column_dimensions['A'].width = 28
for i in range(len(params)):
    ws2.column_dimensions[get_column_letter(2 + i)].width = 25

# ============ SHEET 3: Вердикты ============
ws3 = wb.create_sheet("Вердикты")

ws3.cell(1, 1, "Вердикты и рекомендации").font = Font(bold=True, size=14)
ws3.merge_cells("A1:C1")

ws3.cell(3, 1, "#").font = Font(bold=True)
ws3.cell(3, 2, "Фреймворк").font = Font(bold=True)
ws3.cell(3, 3, "Полезность").font = Font(bold=True)
ws3.cell(3, 4, "Вердикт").font = Font(bold=True)

for si, svc in enumerate(services):
    row = 4 + si
    medals = ["🥇", "🥈", "🥉", "4", "5", "6", "7"]
    ws3.cell(row, 1, medals[min(si, 6)]).border = thin
    ws3.cell(row, 2, svc["name"]).font = Font(bold=True)
    ws3.cell(row, 2).border = thin
    ws3.cell(row, 3, round(svc["_usefulness"], 2)).border = thin
    ws3.cell(row, 4, svc["verdict"]).alignment = wrap
    ws3.cell(row, 4).border = thin

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 60

# ============ SHEET 4: Источники ============
ws4 = wb.create_sheet("Источники")

ws4.cell(1, 1, "Источники информации").font = Font(bold=True, size=14)
ws4.merge_cells("A1:C1")

ws4.cell(3, 1, "Фреймворк").font = Font(bold=True)
ws4.cell(3, 2, "Ссылка").font = Font(bold=True)

row = 4
for svc in services:
    ws4.cell(row, 1, svc["name"]).font = Font(bold=True)
    ws4.cell(row, 1).border = thin
    ws4.cell(row, 1).alignment = Alignment(vertical='top')
    for si, src in enumerate(svc.get("sources", [svc["link"]])):
        cell = ws4.cell(row + si, 2, src)
        cell.font = Font(color="0563C1", size=9)
        cell.border = thin
        cell.alignment = wrap_left
    row += max(len(svc.get("sources", [svc["link"]])), 1)

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 80

# Save
wb.save(output_path)
print(f"Saved: {output_path}")

# Print ranking
print("\n" + "="*60)
print("РЕЙТИНГ ФРЕЙМВОРКОВ (по полезности)")
print("="*60)
for si, svc in enumerate(services):
    medals = ["🥇", "🥈", "🥉", " 4", " 5", " 6", " 7"]
    u = svc["_usefulness"]
    print(f"{medals[min(si,6)]}  {svc['name']:<30} {u:.2f}  {svc['verdict'][:50]}")
